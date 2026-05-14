"""
Excel data access layer — reads and writes the "Data" sheet.
All rows are kept as plain dicts keyed by the exact column header strings.
"""

from __future__ import annotations

import os
from datetime import datetime, date
from typing import Optional

import openpyxl
from openpyxl import load_workbook

SHEET_NAME = "Data"

ALL_COLUMNS = [
    "Platform",
    "ODM",
    "GBU",
    "GTK Supplier",
    "Sub-Category",
    "GTK \nLiability $",
    "Actual GTK \nLiability $",
    "ESR need (Y/N)",
    "Status",
    "DM #",
    "PL",
    "Rebate Initiative %",
    "Actual Payment",
    "Saving",
    "Payment Received Date",
    "Payment Received Quarter",
    "Update Date",
]

DISPLAY_COLUMNS = [
    "Sub-Category",
    "Platform",
    "GTK Supplier",
    "Actual Payment",
    "Status",
    "Payment Received Date",
    "PL",
    "Payment Received Quarter",
    "Update Date",
]


def _quarter_label(d: date) -> str:
    """Return FY quarter string e.g. 'FY26 Q1' for a given date.
    Q1 = Nov, Dec, Jan  (FY = year of Jan)
    Q2 = Feb, Mar, Apr
    Q3 = May, Jun, Jul
    Q4 = Aug, Sep, Oct
    """
    m = d.month
    y = d.year
    if m in (11, 12):
        fy = y + 1
        q = 1
    elif m == 1:
        fy = y
        q = 1
    elif m in (2, 3, 4):
        fy = y
        q = 2
    elif m in (5, 6, 7):
        fy = y
        q = 3
    else:  # 8, 9, 10
        fy = y
        q = 4
    return f"FY{str(fy)[-2:]} Q{q}"


def compute_derived(row: dict) -> dict:
    """Fill in Actual Payment, Saving, Payment Received Quarter from raw fields.
    Rebate Initiative % is stored as a display value (e.g. 10 means 10%).
    """
    try:
        gtk_liability = float(row.get("Actual GTK \nLiability $") or 0)
        rebate_raw = row.get("Rebate Initiative %")
        if rebate_raw is None or rebate_raw == "":
            rebate_pct = 0.0
        else:
            val = float(rebate_raw)
            # Accept both 0.10 (already decimal) and 10 (percent)
            rebate_pct = val if val <= 1.0 else val / 100.0
        actual_payment = gtk_liability * (1 - rebate_pct)
        row["Actual Payment"] = round(actual_payment, 1)
    except (ValueError, TypeError):
        row["Actual Payment"] = None

    try:
        gtk_orig = float(row.get("GTK \nLiability $") or 0)
        row["Saving"] = round(gtk_orig - (row["Actual Payment"] or 0), 1)
    except (ValueError, TypeError):
        row["Saving"] = None

    # ESR need: auto-calculate based on Actual Payment threshold
    actual = row.get("Actual Payment")
    if actual is None:
        row["ESR need (Y/N)"] = ""
    else:
        try:
            row["ESR need (Y/N)"] = "Y" if float(actual) > 500000 else "N"
        except (ValueError, TypeError):
            row["ESR need (Y/N)"] = ""

    prd = row.get("Payment Received Date")
    if prd:
        if isinstance(prd, str):
            try:
                prd = datetime.strptime(prd, "%Y-%m-%d").date()
            except ValueError:
                prd = None
        if isinstance(prd, datetime):
            prd = prd.date()
        if prd:
            row["Payment Received Quarter"] = _quarter_label(prd)
        else:
            row["Payment Received Quarter"] = None
    else:
        row["Payment Received Quarter"] = None

    return row


class ExcelDataStore:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self._rows: list[dict] = []
        self._load()

    # ------------------------------------------------------------------ read --
    def _load(self):
        if not os.path.exists(self.filepath):
            self._rows = []
            return
        wb = load_workbook(self.filepath, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            self._rows = []
            return
        ws = wb[SHEET_NAME]
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in excel_row):
                continue
            row = {}
            for i, h in enumerate(headers):
                if i < len(excel_row):
                    row[h] = excel_row[i]
                else:
                    row[h] = None
            rows.append(row)
        self._rows = rows

    def get_rows(self) -> list[dict]:
        return list(self._rows)

    # ----------------------------------------------------------------- write --
    def append_row(self, row: dict):
        row = compute_derived(row)
        row["Update Date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._rows.append(row)
        self._save()

    def update_row(self, index: int, row: dict):
        row = compute_derived(row)
        row["Update Date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._rows.pop(index)   # remove from original position
        self._rows.append(row)  # re-insert at the bottom
        self._save()

    def delete_row(self, index: int):
        self._rows.pop(index)
        self._save()

    def bulk_append_rows(self, rows: list[dict]) -> int:
        """Compute derived fields for each row and append all at once (single save)."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        count = 0
        for row in rows:
            row = compute_derived(row)
            row["Update Date"] = now
            self._rows.append(row)
            count += 1
        if count:
            self._save()
        return count

    def _save(self):
        if os.path.exists(self.filepath):
            wb = load_workbook(self.filepath)
        else:
            wb = openpyxl.Workbook()
            # remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]

        ws = wb.create_sheet(SHEET_NAME)
        ws.append(ALL_COLUMNS)
        for row in self._rows:
            out = []
            for c in ALL_COLUMNS:
                v = row.get(c)
                if c == "Rebate Initiative %" and v is not None and v != "":
                    try:
                        fv = float(v)
                        # Normalise to decimal fraction for Excel
                        v = fv / 100.0 if fv > 1.0 else fv
                    except (ValueError, TypeError):
                        pass
                out.append(v)
            ws.append(out)

        wb.save(self.filepath)
