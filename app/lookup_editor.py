"""
Lookup Editor — lets user manage the dropdown option lists stored in lookups.json.
"""

from __future__ import annotations

import json
import os
import tkinter as tk
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import load_workbook

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOOKUPS_PATH = os.path.join(ROOT_DIR, "lookups.json")

LOOKUP_KEYS = ["ODM", "GBU", "GTK Supplier", "Sub-Category", "ESR need (Y/N)", "Status"]


def _load_lookups() -> dict:
    if os.path.exists(LOOKUPS_PATH):
        with open(LOOKUPS_PATH, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return {k: [] for k in LOOKUP_KEYS}


def _save_lookups(data: dict):
    with open(LOOKUPS_PATH, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)


class LookupEditorDialog(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Manage Dropdown Options")
        self.geometry("640x520")
        self.resizable(True, True)
        self.grab_set()

        self._data = _load_lookups()
        self._selected_key = tk.StringVar(value=LOOKUP_KEYS[0])
        self._build_ui()
        self._refresh_list()

    def _build_ui(self):
        # Left panel — category selector
        left = ctk.CTkFrame(self, width=180)
        left.pack(side="left", fill="y", padx=(10, 0), pady=10)
        ctk.CTkLabel(left, text="Category", font=ctk.CTkFont(weight="bold")).pack(pady=(8, 4))
        for key in LOOKUP_KEYS:
            btn = ctk.CTkRadioButton(left, text=key, variable=self._selected_key,
                                     value=key, command=self._refresh_list)
            btn.pack(anchor="w", padx=10, pady=2)

        # Right panel — item list + add/remove
        right = ctk.CTkFrame(self)
        right.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        ctk.CTkLabel(right, text="Options (A-Z sorted)", font=ctk.CTkFont(weight="bold")).pack(pady=(8, 4))

        list_frame = tk.Frame(right, bg="#2b2b2b")
        list_frame.pack(fill="both", expand=True, padx=8)

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")
        self._listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                                   bg="#333333", fg="white", selectbackground="#1f538d",
                                   font=("Arial", 11), relief="flat", bd=0)
        self._listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self._listbox.yview)

        add_frame = ctk.CTkFrame(right, fg_color="transparent")
        add_frame.pack(fill="x", padx=8, pady=(6, 0))
        self._new_entry = ctk.CTkEntry(add_frame, placeholder_text="New option...", width=220)
        self._new_entry.pack(side="left", expand=True, fill="x", padx=(0, 6))
        ctk.CTkButton(add_frame, text="Add", width=70, command=self._add_item).pack(side="left", padx=(0, 6))
        ctk.CTkButton(add_frame, text="Import Excel", width=105,
                      fg_color="gray35", hover_color="gray25",
                      command=self._import_excel).pack(side="left")

        btn_row = ctk.CTkFrame(right, fg_color="transparent")
        btn_row.pack(fill="x", padx=8, pady=6)
        ctk.CTkButton(btn_row, text="Remove Selected", fg_color="gray40",
                      hover_color="gray30", command=self._remove_item).pack(side="left")
        ctk.CTkButton(btn_row, text="Save & Close", command=self._save_close).pack(side="right")

    def _refresh_list(self):
        self._listbox.delete(0, tk.END)
        key = self._selected_key.get()
        for item in sorted(self._data.get(key, [])):
            self._listbox.insert(tk.END, item)

    def _add_item(self):
        key = self._selected_key.get()
        val = self._new_entry.get().strip()
        if not val:
            return
        lst = self._data.setdefault(key, [])
        if val not in lst:
            lst.append(val)
        self._new_entry.delete(0, tk.END)
        self._refresh_list()

    def _remove_item(self):
        key = self._selected_key.get()
        sel = self._listbox.curselection()
        if not sel:
            return
        item = self._listbox.get(sel[0])
        if item in self._data.get(key, []):
            self._data[key].remove(item)
        self._refresh_list()

    def _import_excel(self):
        """Bulk-import options from the first column (no header) of an Excel file."""
        path = filedialog.askopenfilename(
            title="Select Excel file to import",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            new_values = []
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                cell = row[0]
                if cell is not None and str(cell).strip():
                    new_values.append(str(cell).strip())
            wb.close()
        except Exception as exc:
            messagebox.showerror("Import Error", f"Could not read file:\n{exc}", parent=self)
            return

        if not new_values:
            messagebox.showinfo("Import", "No values found in the first column.", parent=self)
            return

        key = self._selected_key.get()
        existing = set(self._data.get(key, []))
        added = 0
        for v in new_values:
            if v not in existing:
                existing.add(v)
                added += 1
        self._data[key] = sorted(existing)
        self._refresh_list()
        messagebox.showinfo(
            "Import Complete",
            f"Added {added} new option(s). Duplicates were skipped.",
            parent=self,
        )

    def _save_close(self):
        _save_lookups(self._data)
        self.destroy()
