"""
Main application window — EO Table Management System.
"""

from __future__ import annotations

import json
import os
import sys
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, ttk
from typing import Optional

import customtkinter as ctk

from app.data_store import DISPLAY_COLUMNS, ExcelDataStore
from app.entry_form import EntryFormDialog
from app.lookup_editor import LookupEditorDialog
from app.pl_mapper import load_pl_map, refresh_pl_map
from app.user_selector import UserSelectorDialog

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

CONFIG_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config.json")

# Phase filter labels and their status sets
_PHASE_LABELS = [
    "GTK Deduction Phase",
    "Contract & DM Phase",
    "Complete",
    "Halt",
    "Unknown",
]
_ALL_PHASES: set[str] = set(_PHASE_LABELS)

COLOR_RED    = "#e84646"   # Confirming Numbers
COLOR_YELLOW = "#d4a017"   # Contract & DM process
COLOR_GREEN  = "#3dbb6e"   # Finish
COLOR_BLUE   = "#5bc8e8"   # Halt

_STATUS_GREEN  = {"Finish"}
_STATUS_YELLOW = {"Wait for Contract Approval", "Wait for Contract Sign", "Wait for DM"}
_STATUS_RED    = {"1st Ver Complete", "2nd Ver Complete"}
_STATUS_BLUE   = {"Halt"}

# Column widths for treeview
COL_WIDTHS = {
    "●": 36,
    "Sub-Category": 140,
    "Platform": 130,
    "GTK Supplier": 155,
    "Actual Payment": 140,
    "Status": 130,
    "DM #": 110,
    "Payment Received Date": 170,
    "Payment Received Quarter": 165,
    "Update Date": 175,
}


def _load_config() -> dict:
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return {}


def _save_config(cfg: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as fh:
        json.dump(cfg, fh, indent=2, ensure_ascii=False)


class MainWindow:
    def __init__(self):
        self._root = ctk.CTk()
        self._root.title("EO Table Management System")
        self._root.geometry("1200x720")
        self._root.minsize(900, 600)
        self._root.protocol("WM_DELETE_WINDOW", self._on_close)

        self._store: Optional[ExcelDataStore] = None
        self._filepath: Optional[str] = None
        self._current_user: Optional[str] = None
        self._pl_map: dict[str, str] = load_pl_map()  # Family -> Product Line 2
        self._sort_col: str = "Update Date"  # default: newest first
        self._sort_asc: bool = False
        self._filter_var = tk.StringVar()
        self._filter_var.trace_add("write", lambda *_: self._refresh_table())
        self._seg_btn: Optional[ctk.CTkButton] = None
        # Column-level filters: col -> set of allowed values (missing key = no filter)
        self._col_filters: dict[str, set] = {}
        # Phase filter: set of active phase labels (all = no filter)
        self._phase_filter: set[str] = set(_PHASE_LABELS)  # default: all selected

        self._segment_active: set = set()  # kept for compat
        self._all_segments: list = []

        self._build_ui()
        # Set initial sort arrow
        self._tree.heading("Update Date", text="Update Date ▼")

    # ------------------------------------------------------------------ UI ---
    def _build_ui(self):
        # ---- Top bar — buttons packed RIGHT first so they're always visible ----
        top = ctk.CTkFrame(self._root, height=52, corner_radius=0)
        top.pack(side="top", fill="x")
        top.pack_propagate(False)

        # Right-side buttons (packed before path label to guarantee visibility)
        ctk.CTkButton(top, text="+ Add Entry", width=105, height=32,
                      font=ctk.CTkFont(size=12), command=self._add_entry).pack(side="right", padx=(4, 12))
        ctk.CTkButton(top, text="Manage Options", width=128, height=32,
                      font=ctk.CTkFont(size=12), fg_color="gray35", hover_color="gray25",
                      command=self._open_lookup_editor).pack(side="right", padx=4)
        ctk.CTkButton(top, text="Import Excel", width=110, height=32,
                      font=ctk.CTkFont(size=12), fg_color="#4a4a8a", hover_color="#35356e",
                      command=self._import_excel).pack(side="right", padx=4)
        ctk.CTkButton(top, text="Connect Data", width=110, height=32,
                      font=ctk.CTkFont(size=12), command=self._browse_file).pack(side="right", padx=4)
        self._refresh_pl_btn = ctk.CTkButton(
            top, text="Refresh PL", width=110, height=32,
            font=ctk.CTkFont(size=12), fg_color="#2a6040", hover_color="#1e4a2e",
            command=self._refresh_pl,
        )
        self._refresh_pl_btn.pack(side="right", padx=4)

        # Left: title + user badge + file name
        ctk.CTkLabel(top, text="EO Table Management",
                     font=ctk.CTkFont(size=15, weight="bold")).pack(side="left", padx=(14, 6))
        self._user_label = ctk.CTkLabel(top, text="", font=ctk.CTkFont(size=12),
                                        text_color="#4db8ff")
        self._user_label.pack(side="left", padx=(0, 8))
        self._path_label = ctk.CTkLabel(top, text="No file loaded",
                                        font=ctk.CTkFont(size=11), text_color="gray60", anchor="w")
        self._path_label.pack(side="left", padx=4, expand=True, fill="x")

        # ---- Filter bar ----
        filter_bar = ctk.CTkFrame(self._root, height=46, corner_radius=0, fg_color="#1e1e1e")
        filter_bar.pack(side="top", fill="x")
        filter_bar.pack_propagate(False)
        ctk.CTkLabel(filter_bar, text="Search:", font=ctk.CTkFont(size=13)).pack(side="left", padx=(14, 4))
        ctk.CTkEntry(filter_bar, textvariable=self._filter_var, width=320, height=32,
                     font=ctk.CTkFont(size=13),
                     placeholder_text="Filter any column...").pack(side="left", padx=(4, 10))

        ctk.CTkLabel(filter_bar, text="Right-click column header to filter",
                     font=ctk.CTkFont(size=11), text_color="gray50").pack(side="left", padx=(0, 4))

        # Phase filter button (right side of filter bar)
        self._phase_btn = ctk.CTkButton(
            filter_bar, text="Phase ▾", width=120, height=32,
            font=ctk.CTkFont(size=12), fg_color="gray35", hover_color="gray25",
            command=self._open_phase_filter,
        )
        self._phase_btn.pack(side="right", padx=(4, 14))

        # ---- Table area ----
        table_frame = tk.Frame(self._root, bg="#1e1e1e")
        table_frame.pack(fill="both", expand=True, padx=8, pady=8)

        # Treeview style
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                        background="#2b2b2b",
                        foreground="white",
                        rowheight=34,
                        fieldbackground="#2b2b2b",
                        font=("Arial", 12))
        style.configure("Custom.Treeview.Heading",
                        background="#1f538d",
                        foreground="white",
                        relief="flat",
                        font=("Arial", 12, "bold"))
        style.map("Custom.Treeview",
                  background=[("selected", "#1f538d")],
                  foreground=[("selected", "white")])
        style.map("Custom.Treeview.Heading",
                  background=[("active", "#174b7a")])

        cols = ["●"] + DISPLAY_COLUMNS
        self._tree = ttk.Treeview(table_frame, columns=cols, show="headings",
                                  style="Custom.Treeview", selectmode="browse")

        # Scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        # Configure columns
        for col in cols:
            w = COL_WIDTHS.get(col, 120)
            self._tree.heading(col, text=col, anchor="center",
                               command=lambda c=col: self._sort_by(c))
            self._tree.column(col, width=w, minwidth=40, stretch=(col != "●"), anchor="center")

        self._col_header_base = {c: c for c in cols}  # track base label for sort arrows

        # Row tags for status indicator colours
        self._tree.tag_configure("red_dot",     foreground=COLOR_RED)
        self._tree.tag_configure("yellow_dot",  foreground=COLOR_YELLOW)
        self._tree.tag_configure("green_dot",   foreground=COLOR_GREEN)
        self._tree.tag_configure("blue_dot",    foreground=COLOR_BLUE)
        self._tree.tag_configure("stripe",      background="#313131")
        self._tree.tag_configure("stripe_red",    background="#313131", foreground=COLOR_RED)
        self._tree.tag_configure("stripe_yellow", background="#313131", foreground=COLOR_YELLOW)
        self._tree.tag_configure("stripe_green",  background="#313131", foreground=COLOR_GREEN)
        self._tree.tag_configure("stripe_blue",   background="#313131", foreground=COLOR_BLUE)

        # Double-click to edit
        self._tree.bind("<Double-1>", self._on_double_click)
        # Right-click on heading → column filter popup
        self._tree.bind("<Button-3>", self._on_header_right_click)
        # Tooltip for ⚠ ESR cells
        self._tip_win: Optional[tk.Toplevel] = None
        self._tree.bind("<Motion>", self._on_tree_motion)
        self._tree.bind("<Leave>", self._hide_tip)

        # ---- Status bar ----
        self._status_var = tk.StringVar(value="Ready")
        status_bar = ctk.CTkLabel(self._root, textvariable=self._status_var,
                                  font=ctk.CTkFont(size=12), text_color="gray60", anchor="w")
        status_bar.pack(side="bottom", fill="x", padx=14, pady=(0, 5))

    # --------------------------------------------------------------- actions --
    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        )
        if path:
            self._load_file(path)

    # ---- column name normaliser for import ----
    _IMPORT_COL_MAP: dict[str, str] = {
        "platform":                    "Platform",
        "odm":                         "ODM",
        "gbu":                         "GBU",
        "gtk supplier":                "GTK Supplier",
        "sub-category":                "Sub-Category",
        "gtk liability $":             "GTK \nLiability $",
        "gtk \nliability $":           "GTK \nLiability $",
        "actual gtk liability $":      "Actual GTK \nLiability $",
        "actual gtk \nliability $":    "Actual GTK \nLiability $",
        "dm #":                        "DM #",
        "pl":                          "PL",
        "rebate initiative %":         "Rebate Initiative %",
        "actual payment":              "Actual Payment",
        "saving":                      "Saving",
        "payment received date":       "Payment Received Date",
        "status":                      "Status",
    }

    def _import_excel(self):
        if self._store is None:
            messagebox.showwarning("No Data File", "Please connect a data file first (Connect Data).")
            return

        path = filedialog.askopenfilename(
            title="Select Import Excel File",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not path:
            return

        try:
            import openpyxl as _openpyxl
            wb = _openpyxl.load_workbook(path, data_only=True)
            ws = wb.active  # use the first/active sheet

            raw_headers = [
                str(c.value).strip() if c.value is not None else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            # Map each source column index → internal column name (or None to skip)
            col_map: list[str | None] = []
            for h in raw_headers:
                normalized = h.lower().replace("\n", " ").replace("  ", " ").strip()
                col_map.append(self._IMPORT_COL_MAP.get(normalized))

            rows_to_import: list[dict] = []
            for excel_row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in excel_row):
                    continue
                row: dict = {}
                for i, internal_col in enumerate(col_map):
                    if internal_col is None:
                        continue
                    val = excel_row[i] if i < len(excel_row) else None
                    row[internal_col] = val
                # skip entirely blank rows
                if not any(row.values()):
                    continue
                rows_to_import.append(row)

            if not rows_to_import:
                messagebox.showinfo("Import Excel", "No data rows found in the selected file.")
                return

            count = self._store.bulk_append_rows(rows_to_import)
            self._refresh_table()
            self._status_var.set(f"Imported {count} rows from: {os.path.basename(path)}")
            messagebox.showinfo("Import Complete", f"Successfully imported {count} rows.")

        except Exception as exc:
            messagebox.showerror("Import Error", f"Failed to import file:\n{exc}")

    def _load_file(self, path: str):
        try:
            self._store = ExcelDataStore(path)
            self._filepath = path
            self._path_label.configure(text=os.path.basename(path))
            cfg = _load_config()
            user_paths = cfg.get("user_paths", {})
            if self._current_user:
                user_paths[self._current_user] = path
                cfg["user_paths"] = user_paths
            _save_config(cfg)
            self._refresh_table()
            n = len(self._store.get_rows())
            self._status_var.set(f"Loaded {n} rows from: {os.path.basename(path)}")
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to load file:\n{exc}")

    def _on_header_right_click(self, event):
        """Open a checkbox filter popup when right-clicking a column heading."""
        region = self._tree.identify_region(event.x, event.y)
        if region != "heading":
            return
        col_id = self._tree.identify_column(event.x)
        col_idx = int(col_id.lstrip("#")) - 1
        cols = ["●"] + DISPLAY_COLUMNS
        if col_idx < 0 or col_idx >= len(cols):
            return
        col = cols[col_idx]
        if col == "●":
            return
        self._open_col_filter(col, event.x_root, event.y_root)

    def _open_col_filter(self, col: str, x: int, y: int):
        """Show a multi-select checkbox filter for the given column."""
        if self._store is None:
            return
        all_values = sorted(
            {str(row.get(col) or "").strip() for row in self._store.get_rows()},
            key=lambda v: ("ÿ" if v == "" else v.lower()),
        )
        if not all_values:
            return

        current = self._col_filters.get(col, set())
        popup = ctk.CTkToplevel(self._root)
        popup.title("")
        popup.overrideredirect(True)
        popup.configure(fg_color="#2b2b2b")
        # Do NOT grab_set — we need outside clicks to register for auto-close

        ctk.CTkLabel(popup, text=f"Filter: {col.replace(chr(10), ' ')}",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(padx=14, pady=(10, 4))
        ctk.CTkFrame(popup, height=1, fg_color="gray35").pack(fill="x", padx=10, pady=2)

        temp_vars: dict[str, tk.BooleanVar] = {}
        for v in all_values:
            temp_vars[v] = tk.BooleanVar(value=(not current) or (v in current))

        all_var = tk.BooleanVar(value=not current)

        def _on_all():
            on = all_var.get()
            for bv in temp_vars.values():
                bv.set(on)

        def _on_item():
            all_var.set(all(bv.get() for bv in temp_vars.values()))

        ctk.CTkCheckBox(popup, text="(Select All)", variable=all_var,
                        font=ctk.CTkFont(size=12, weight="bold"),
                        command=_on_all).pack(anchor="w", padx=14, pady=3)

        scroll_h = min(len(all_values) * 30 + 10, 260)
        scroll = ctk.CTkScrollableFrame(popup, fg_color="transparent",
                                        width=230, height=scroll_h)
        scroll.pack(padx=6, pady=2)
        for v in all_values:
            ctk.CTkCheckBox(scroll, text=(v if v else "(blank)"),
                            variable=temp_vars[v], font=ctk.CTkFont(size=12),
                            command=_on_item).pack(anchor="w", pady=2)

        ctk.CTkFrame(popup, height=1, fg_color="gray35").pack(fill="x", padx=10, pady=2)
        btn_row = ctk.CTkFrame(popup, fg_color="transparent")
        btn_row.pack(fill="x", padx=14, pady=(4, 10))

        def _apply():
            chosen = {v for v, bv in temp_vars.items() if bv.get()}
            if len(chosen) == len(all_values):
                self._col_filters.pop(col, None)
            else:
                self._col_filters[col] = chosen
            self._update_heading_text(col)
            self._refresh_table()
            self._save_user_filters()
            popup.destroy()

        def _clear():
            self._col_filters.pop(col, None)
            self._update_heading_text(col)
            self._refresh_table()
            self._save_user_filters()
            popup.destroy()

        ctk.CTkButton(btn_row, text="Apply", width=90, height=30,
                      font=ctk.CTkFont(size=12), command=_apply).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="Clear", width=80, height=30,
                      font=ctk.CTkFont(size=12), fg_color="gray40", hover_color="gray30",
                      command=_clear).pack(side="left")

        popup.update_idletasks()
        popup.geometry(f"+{x}+{y}")

        def _close_if_outside(event):
            if not popup.winfo_exists():
                return
            px, py = popup.winfo_rootx(), popup.winfo_rooty()
            pw, ph = popup.winfo_width(), popup.winfo_height()
            if not (px <= event.x_root <= px + pw and py <= event.y_root <= py + ph):
                popup.destroy()

        # bind_all fires for every widget click in the whole app
        popup.bind_all("<Button-1>", _close_if_outside, add="+")
        popup.bind("<Destroy>", lambda e: popup.unbind_all("<Button-1>"))

    def _update_heading_text(self, col: str):
        """Rebuild heading label with sort arrow and/or filter indicator."""
        if not hasattr(self, "_col_header_base"):
            return
        base = self._col_header_base.get(col, col)
        text = base
        if self._sort_col == col:
            text += " ▲" if self._sort_asc else " ▼"
        if col in self._col_filters:
            text += " ◆"
        self._tree.heading(col, text=text)

    def _rebuild_segment_bar(self):
        pass  # removed — replaced by per-column filter

    def _update_seg_btn_label(self):
        pass

    def _open_segment_dropdown(self):
        pass

    def _toggle_all_segments(self):
        pass

    def _toggle_segment(self, seg):
        pass

    def _update_segment_button_styles(self):
        pass

    def _save_segment_prefs(self):
        pass

    def _row_phase(self, row: dict) -> str:
        """Return the phase label for a row based on its Status."""
        status = str(row.get("Status") or "").strip()
        if status in _STATUS_RED:
            return "GTK Deduction Phase"
        if status in _STATUS_YELLOW:
            return "Contract & DM Phase"
        if status in _STATUS_GREEN:
            return "Complete"
        if status in _STATUS_BLUE:
            return "Halt"
        return "Unknown"

    def _update_phase_btn_text(self):
        if self._phase_filter == _ALL_PHASES:
            self._phase_btn.configure(text="Phase ▾")
        else:
            self._phase_btn.configure(text=f"Phase ▾ ({len(self._phase_filter)})")

    def _open_phase_filter(self):
        popup = ctk.CTkToplevel(self._root)
        popup.title("")
        popup.overrideredirect(True)
        popup.configure(fg_color="#2b2b2b")

        ctk.CTkLabel(popup, text="Filter by Phase",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(padx=14, pady=(10, 4))
        ctk.CTkFrame(popup, height=1, fg_color="gray35").pack(fill="x", padx=10, pady=2)

        _PHASE_COLORS = {
            "GTK Deduction Phase": COLOR_RED,
            "Contract & DM Phase": COLOR_YELLOW,
            "Complete":            COLOR_GREEN,
            "Halt":                COLOR_BLUE,
            "Unknown":             "gray70",
        }

        temp_vars: dict[str, tk.BooleanVar] = {}
        for label in _PHASE_LABELS:
            temp_vars[label] = tk.BooleanVar(value=(label in self._phase_filter))

        all_var = tk.BooleanVar(value=(self._phase_filter == _ALL_PHASES))

        def _on_all():
            on = all_var.get()
            for bv in temp_vars.values():
                bv.set(on)

        def _on_item():
            all_var.set(all(bv.get() for bv in temp_vars.values()))

        ctk.CTkCheckBox(popup, text="(Select All)", variable=all_var,
                        font=ctk.CTkFont(size=12),
                        command=_on_all).pack(anchor="w", padx=14, pady=(4, 2))
        ctk.CTkFrame(popup, height=1, fg_color="gray35").pack(fill="x", padx=10, pady=2)

        for label in _PHASE_LABELS:
            color = _PHASE_COLORS[label]
            ctk.CTkCheckBox(popup, text=label, variable=temp_vars[label],
                            font=ctk.CTkFont(size=12), text_color=color,
                            command=_on_item).pack(anchor="w", padx=14, pady=2)

        btn_row = ctk.CTkFrame(popup, fg_color="transparent")
        btn_row.pack(fill="x", padx=14, pady=(8, 10))

        def _apply():
            chosen = {lb for lb, bv in temp_vars.items() if bv.get()}
            self._phase_filter = chosen if chosen else _ALL_PHASES
            self._update_phase_btn_text()
            self._refresh_table()
            self._save_user_filters()
            popup.destroy()

        def _clear():
            self._phase_filter = set(_PHASE_LABELS)
            self._update_phase_btn_text()
            self._refresh_table()
            self._save_user_filters()
            popup.destroy()

        ctk.CTkButton(btn_row, text="Apply", width=90, height=30,
                      font=ctk.CTkFont(size=12), command=_apply).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btn_row, text="Clear", width=80, height=30,
                      font=ctk.CTkFont(size=12), fg_color="gray40", hover_color="gray30",
                      command=_clear).pack(side="left")

        popup.update_idletasks()
        bx = self._phase_btn.winfo_rootx()
        by = self._phase_btn.winfo_rooty() + self._phase_btn.winfo_height() + 4
        popup.geometry(f"+{bx}+{by}")

        def _close_if_outside(event):
            if not popup.winfo_exists():
                return
            px, py = popup.winfo_rootx(), popup.winfo_rooty()
            pw, ph = popup.winfo_width(), popup.winfo_height()
            if not (px <= event.x_root <= px + pw and py <= event.y_root <= py + ph):
                popup.destroy()
                self._root.unbind_all("<Button-1>")
        popup.after(100, lambda: self._root.bind_all("<Button-1>", _close_if_outside))

    def _refresh_table(self):
        if self._store is None:
            return
        rows = self._store.get_rows()

        # Apply phase filter
        if self._phase_filter != _ALL_PHASES:
            rows = [r for r in rows if self._row_phase(r) in self._phase_filter]

        # Apply column-level filters
        for _col, _allowed in self._col_filters.items():
            if _allowed:
                rows = [r for r in rows
                        if str(r.get(_col) or "").strip() in _allowed]

        # Apply text filter
        flt = self._filter_var.get().lower().strip()
        if flt:
            rows = [r for r in rows if any(flt in str(v).lower() for v in r.values())]

        # Apply sort
        if self._sort_col == "●":
            # Sort by payment status: ascending = red (no payment) first
            rows = sorted(rows,
                          key=lambda r: (1 if bool(r.get("Payment Received Date")) else 0),
                          reverse=not self._sort_asc)
        elif self._sort_col:
            def _key(r):
                v = r.get(self._sort_col)
                if v is None:
                    return ("", )
                return (str(v).lower(), )
            rows = sorted(rows, key=_key, reverse=not self._sort_asc)

        # Redraw — use after() to avoid blocking
        self._root.after(0, lambda: self._populate_tree(rows))

    def _populate_tree(self, rows: list[dict]):
        self._tree.delete(*self._tree.get_children())
        for i, row in enumerate(rows):
            stripe = (i % 2 == 1)

            values = ["●"] + [self._fmt_col(c, row.get(c)) for c in DISPLAY_COLUMNS]

            status = str(row.get("Status") or "").strip()
            if status in _STATUS_GREEN:
                tag = "stripe_green" if stripe else "green_dot"
            elif status in _STATUS_YELLOW:
                tag = "stripe_yellow" if stripe else "yellow_dot"
            elif status in _STATUS_RED:
                tag = "stripe_red" if stripe else "red_dot"
            elif status in _STATUS_BLUE:
                tag = "stripe_blue" if stripe else "blue_dot"
            else:
                tag = "stripe" if stripe else ""

            self._tree.insert("", "end", iid=str(id(row)) + str(i),
                               values=values, tags=(tag,))

        # Store rows list for edit lookup
        self._current_rows = rows

    def _fmt(self, val) -> str:
        if val is None:
            return ""
        return str(val)

    def _fmt_col(self, col: str, val) -> str:
        if val is None or val == "":
            return ""
        if col == "Actual Payment":
            try:
                amount = float(val)
                formatted = f"${amount:,.1f}"
                if amount > 500_000:
                    formatted = "⚠ " + formatted
                return formatted
            except (ValueError, TypeError):
                return str(val)
        return str(val)

    def _on_tree_motion(self, event: tk.Event):
        """Show 'ESR Needed' tooltip when hovering over a ⚠ Actual Payment cell."""
        region = self._tree.identify_region(event.x, event.y)
        if region != "cell":
            self._hide_tip()
            return
        col_id = self._tree.identify_column(event.x)
        try:
            col_index = int(col_id.lstrip("#")) - 1
        except ValueError:
            self._hide_tip()
            return
        cols = ["●"] + DISPLAY_COLUMNS
        if col_index < 0 or col_index >= len(cols) or cols[col_index] != "Actual Payment":
            self._hide_tip()
            return
        row_id = self._tree.identify_row(event.y)
        if not row_id:
            self._hide_tip()
            return
        cell_val = self._tree.set(row_id, col_id)
        if not cell_val.startswith("⚠"):
            self._hide_tip()
            return
        # Show tooltip
        x = event.x_root + 14
        y = event.y_root + 14
        if self._tip_win:
            self._tip_win.geometry(f"+{x}+{y}")
            return
        self._tip_win = tk.Toplevel(self._root)
        self._tip_win.wm_overrideredirect(True)
        self._tip_win.wm_geometry(f"+{x}+{y}")
        lbl = tk.Label(self._tip_win, text="ESR Needed",
                       background="#ffdd57", foreground="#1a1a1a",
                       font=("Arial", 10, "bold"), relief="solid",
                       borderwidth=1, padx=6, pady=3)
        lbl.pack()

    def _hide_tip(self, _event=None):
        if self._tip_win:
            self._tip_win.destroy()
            self._tip_win = None

    def _sort_by(self, col: str):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            prev = self._sort_col
            self._sort_col = col
            self._sort_asc = True
            if prev and hasattr(self, "_col_header_base"):
                self._update_heading_text(prev)
        self._update_heading_text(col)
        self._refresh_table()

    def _add_entry(self):
        if self._store is None:
            messagebox.showwarning("No file", "Please open an Excel file first.")
            return
        dlg = EntryFormDialog(self._root, title="Add Entry", store=self._store,
                              pl_map=self._pl_map)
        self._root.wait_window(dlg)
        result = dlg.get_result()
        if result:
            self._store.append_row(result)
            self._refresh_table()
            self._status_var.set("Entry added.")

    def _on_double_click(self, event):
        item = self._tree.focus()
        if not item:
            return
        idx = self._tree.index(item)
        if not hasattr(self, "_current_rows") or idx >= len(self._current_rows):
            return
        existing = self._current_rows[idx]
        # Find actual index in store
        all_rows = self._store.get_rows()
        try:
            store_idx = all_rows.index(existing)
        except ValueError:
            store_idx = None

        dlg = EntryFormDialog(self._root, existing_row=existing, title="Edit Entry",
                              store=self._store, pl_map=self._pl_map)
        self._root.wait_window(dlg)
        if dlg.is_deleted() and store_idx is not None:
            self._store.delete_row(store_idx)
            self._refresh_table()
            self._status_var.set("Entry deleted.")
        elif dlg.get_result() is not None and store_idx is not None:
            self._store.update_row(store_idx, dlg.get_result())
            self._refresh_table()
            self._status_var.set("Entry updated.")

    def _open_lookup_editor(self):
        dlg = LookupEditorDialog(self._root)
        self._root.wait_window(dlg)

    # ---------------------------------------------------------- PL refresh ---
    def _refresh_pl(self):
        """Run PL map refresh in a background thread to keep UI responsive."""
        import threading
        self._refresh_pl_btn.configure(state="disabled", text="Refreshing…")
        self._status_var.set("Refreshing PL map…")

        def _run():
            try:
                n = refresh_pl_map()
                self._root.after(0, lambda: self._on_pl_refresh_done(n, None))
            except Exception as exc:
                err = str(exc)
                self._root.after(0, lambda: self._on_pl_refresh_done(None, err))

        threading.Thread(target=_run, daemon=True).start()

    def _on_pl_refresh_done(self, count, error):
        self._refresh_pl_btn.configure(state="normal", text="Refresh PL")
        if error:
            self._status_var.set("Refresh PL failed.")
            self._root.lift()
            self._root.focus_force()
            messagebox.showerror("Refresh PL Failed", error, parent=self._root)
        else:
            self._pl_map = load_pl_map()
            self._status_var.set(f"PL map refreshed — {count} unique entries.")
            self._root.lift()
            self._root.focus_force()
            messagebox.showinfo(
                "Refresh PL Complete",
                f"PL map refreshed successfully.\n{count} unique Platform → PL entries.",
                parent=self._root,
            )

    def _save_user_filters(self):
        """Persist current col_filters and phase_filter for the logged-in user to config.json."""
        if not self._current_user:
            return
        cfg = _load_config()
        user_filters = cfg.get("user_filters", {})
        user_filters[self._current_user] = {
            col: sorted(vals) for col, vals in self._col_filters.items()
        }
        cfg["user_filters"] = user_filters
        # Save phase filter (store as list; missing = all selected)
        user_phase = cfg.get("user_phase_filters", {})
        if self._phase_filter == _ALL_PHASES:
            user_phase.pop(self._current_user, None)
        else:
            user_phase[self._current_user] = sorted(self._phase_filter)
        cfg["user_phase_filters"] = user_phase
        _save_config(cfg)

    def _load_user_filters(self, user: str):
        """Restore col_filters and phase_filter for the given user from config.json."""
        cfg = _load_config()
        saved = cfg.get("user_filters", {}).get(user, {})
        self._col_filters = {col: set(vals) for col, vals in saved.items()}
        for col in self._col_filters:
            self._update_heading_text(col)
        # Restore phase filter
        saved_phase = cfg.get("user_phase_filters", {}).get(user)
        if saved_phase is not None:
            self._phase_filter = set(saved_phase) & _ALL_PHASES
        else:
            self._phase_filter = set(_PHASE_LABELS)
        self._update_phase_btn_text()

    def _select_user_at_startup(self):
        """Show user selector and load the user's saved Excel path."""
        cfg = _load_config()
        users = cfg.get("users", [])
        last_user = cfg.get("last_user")

        dlg = UserSelectorDialog(self._root, users=users, last_user=last_user)
        self._root.wait_window(dlg)

        selected = dlg.get_selected()
        if not selected:
            self._root.destroy()
            sys.exit(0)

        all_users = dlg.get_all_users()
        cfg["users"] = all_users
        cfg["last_user"] = selected
        _save_config(cfg)

        self._current_user = selected
        self._user_label.configure(text=f"👤 {selected}")
        self._root.title(f"EO Table Management — {selected}")

        # Restore this user's column filters
        self._load_user_filters(selected)

        # Load user's remembered Excel path
        user_paths = cfg.get("user_paths", {})
        path = user_paths.get(selected, "")
        if path and os.path.exists(path):
            self._load_file(path)  # _load_file calls _rebuild_segment_bar
        else:
            self._rebuild_segment_bar()
            self._status_var.set(f"Welcome, {selected}! Open an Excel file to get started.")

    def _on_close(self):
        self._root.destroy()
        sys.exit(0)

    def run(self):
        self._root.after(80, self._select_user_at_startup)
        self._root.mainloop()
