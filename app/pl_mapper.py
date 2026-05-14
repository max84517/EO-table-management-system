"""
PL Mapper
---------
Reads all Excel files from  data/PL Source/  (sheet "ODM L10"),
processes the "Family" column, and writes a deduplicated
data/PL output/PL map.xlsx with columns: Origin | Family | Product Line 2.

Processing rules for the Family value
  1. Remove any FYXXQX substring (case-insensitive, e.g. FY24Q2, fy23q1)
     including one optional space immediately before it.
  2. Split on "/" → produces one row per part.
  3. For each part: if text after the first space is NOT "Early closed", delete it.
  4. If "Early closed" appeared anywhere in the original (post-step-1) value,
     append " Early closed" to every split part.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
import pyxlsb

# ── path resolution (works both in dev mode and frozen exe) ──────────────────
if getattr(sys, "frozen", False):
    _ROOT = Path(sys.executable).parent
else:
    _ROOT = Path(__file__).parent.parent

PL_SOURCE_DIR = _ROOT / "data" / "PL Source"
PL_OUTPUT_DIR = _ROOT / "data" / "PL output"
PL_MAP_FILE   = PL_OUTPUT_DIR / "PL map.xlsx"

SHEET_NAME = "ODM L10"
COL_FAMILY = "Family"
COL_PL2    = "Product Line 2"

# Regex: optional space + FY + 2 digits + Q + 1 digit (case-insensitive)
_FY_PAT = re.compile(r"\s*FY\d{2}Q\d\b", re.IGNORECASE)
_EC_PAT = re.compile(r"early\s+closed", re.IGNORECASE)


# ── Family processing ─────────────────────────────────────────────────────────

def _process_family(raw: str) -> list[str]:
    """
    Return one or more cleaned Family strings from a raw cell value.
    See module docstring for rules.
    """
    if not isinstance(raw, str) or not raw.strip():
        return [str(raw).strip() if raw else ""]

    # Step 1: strip FYXXQX
    cleaned = _FY_PAT.sub("", raw).strip()

    # Step 2: split on "/"
    parts = [p.strip() for p in cleaned.split("/")]

    # Step 3 & 4: detect "Early closed" per-part, strip unwanted suffixes
    has_early_closed = False
    clean_parts: list[str] = []

    for part in parts:
        space_idx = part.find(" ")
        if space_idx == -1:
            clean_parts.append(part)
        else:
            rest = part[space_idx + 1:].strip()
            if _EC_PAT.match(rest):          # rest IS "Early closed"
                has_early_closed = True
                clean_parts.append(part[:space_idx])
            else:
                # Some other suffix → delete it
                clean_parts.append(part[:space_idx])

    # Re-attach "Early closed" to every part if it was present
    return [
        f"{p} Early closed" if has_early_closed else p
        for p in clean_parts
    ]


# ── per-file reader ───────────────────────────────────────────────────────────

def _read_source_file(path: Path) -> list[dict]:
    """Return processed rows from one Excel source file (.xlsx/.xls or .xlsb)."""
    if path.suffix.lower() == ".xlsb":
        return _read_xlsb(path)
    return _read_xlsx(path)


def _find_sheet_name(available: list, target: str) -> str | None:
    """Return the actual sheet name matching target (case-insensitive, stripped)."""
    t = target.strip().lower()
    for name in available:
        if str(name).strip().lower() == t:
            return name
    return None


def _read_xlsx(path: Path) -> list[dict]:
    """Reader for .xlsx / .xls files via openpyxl."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise RuntimeError(f"Cannot open {path.name}: {exc}") from exc

    actual_sheet = _find_sheet_name(wb.sheetnames, SHEET_NAME)
    if actual_sheet is None:
        wb.close()
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in {path.name}")

    ws = wb[actual_sheet]
    header_map: dict[str, int] = {}
    data_rows = ws.iter_rows(values_only=True)

    for raw_row in data_rows:
        for i, cell in enumerate(raw_row):
            if cell is not None:
                header_map[str(cell).strip()] = i
        if COL_FAMILY in header_map and COL_PL2 in header_map:
            break
    else:
        wb.close()
        raise RuntimeError(
            f"Columns '{COL_FAMILY}' / '{COL_PL2}' not found in "
            f"{path.name} (sheet '{SHEET_NAME}')"
        )

    fam_i = header_map[COL_FAMILY]
    pl2_i = header_map[COL_PL2]
    results: list[dict] = []
    for raw_row in data_rows:
        family_raw = raw_row[fam_i] if fam_i < len(raw_row) else None
        pl2_val    = raw_row[pl2_i] if pl2_i < len(raw_row) else None
        if family_raw is None or str(family_raw).strip() == "":
            continue
        origin  = str(family_raw).strip()
        pl2_str = str(pl2_val).strip() if pl2_val is not None else ""
        for fam in _process_family(origin):
            if fam:
                results.append({"Origin": origin, COL_FAMILY: fam, COL_PL2: pl2_str})
    wb.close()
    return results


def _read_xlsb(path: Path) -> list[dict]:
    """Reader for .xlsb files via pyxlsb."""
    try:
        wb = pyxlsb.open_workbook(str(path))
    except Exception as exc:
        raise RuntimeError(f"Cannot open {path.name}: {exc}") from exc

    actual_sheet = _find_sheet_name(wb.sheets, SHEET_NAME)
    if actual_sheet is None:
        wb.close()
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in {path.name}")

    results: list[dict] = []
    with wb.get_sheet(actual_sheet) as ws:
        rows_iter = ws.rows()
        # Locate header row
        header_map: dict[str, int] = {}
        for raw_row in rows_iter:
            for i, cell in enumerate(raw_row):
                if cell.v is not None:
                    header_map[str(cell.v).strip()] = i
            if COL_FAMILY in header_map and COL_PL2 in header_map:
                break
        else:
            wb.close()
            raise RuntimeError(
                f"Columns '{COL_FAMILY}' / '{COL_PL2}' not found in "
                f"{path.name} (sheet '{SHEET_NAME}')"
            )

        fam_i = header_map[COL_FAMILY]
        pl2_i = header_map[COL_PL2]
        for raw_row in rows_iter:
            fam_cell = raw_row[fam_i] if fam_i < len(raw_row) else None
            pl2_cell = raw_row[pl2_i] if pl2_i < len(raw_row) else None
            family_raw = fam_cell.v if fam_cell else None
            pl2_val    = pl2_cell.v if pl2_cell else None
            if family_raw is None or str(family_raw).strip() == "":
                continue
            origin  = str(family_raw).strip()
            pl2_str = str(pl2_val).strip() if pl2_val is not None else ""
            for fam in _process_family(origin):
                if fam:
                    results.append({"Origin": origin, COL_FAMILY: fam, COL_PL2: pl2_str})

    wb.close()
    return results


# ── public API ────────────────────────────────────────────────────────────────

def refresh_pl_map() -> int:
    """
    Read all .xlsx / .xls files from PL_SOURCE_DIR, process them,
    deduplicate on (Family, Product Line 2), and write PL_MAP_FILE.

    Returns the number of unique rows written.
    Raises RuntimeError if no files found or all files fail to parse.
    """
    PL_SOURCE_DIR.mkdir(parents=True, exist_ok=True)
    PL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(
        list(PL_SOURCE_DIR.glob("*.xlsx"))
        + list(PL_SOURCE_DIR.glob("*.xls"))
        + list(PL_SOURCE_DIR.glob("*.xlsb"))
    )
    if not xlsx_files:
        raise RuntimeError(f"No Excel files found in:\n{PL_SOURCE_DIR}")

    all_rows: list[dict] = []
    errors:   list[str]  = []

    for f in xlsx_files:
        try:
            all_rows.extend(_read_source_file(f))
        except RuntimeError as exc:
            errors.append(str(exc))

    if not all_rows:
        msg = "No data could be read."
        if errors:
            msg += "\n\nErrors:\n" + "\n".join(errors)
        raise RuntimeError(msg)

    # Deduplicate on (Family, Product Line 2)
    seen:        set[tuple]  = set()
    unique_rows: list[dict]  = []
    for row in all_rows:
        key = (row[COL_FAMILY], row[COL_PL2])
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)

    # Write output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "PL Map"
    ws_out.append(["Origin", COL_FAMILY, COL_PL2])
    for row in unique_rows:
        ws_out.append([row["Origin"], row[COL_FAMILY], row[COL_PL2]])

    wb_out.save(PL_MAP_FILE)
    return len(unique_rows)


def load_pl_map() -> dict[str, str]:
    """
    Load PL map.xlsx and return  { Family: Product_Line_2 }.
    Returns an empty dict if the file does not yet exist or cannot be read.
    """
    if not PL_MAP_FILE.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(PL_MAP_FILE, data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if headers is None:
            wb.close()
            return {}
        h = [str(c).strip() if c else "" for c in headers]
        try:
            fam_i = h.index(COL_FAMILY)
            pl2_i = h.index(COL_PL2)
        except ValueError:
            wb.close()
            return {}
        result: dict[str, str] = {}
        for row in rows_iter:
            fam = row[fam_i] if fam_i < len(row) else None
            pl2 = row[pl2_i] if pl2_i < len(row) else None
            if fam and str(fam).strip():
                # Store lowercase key for case-insensitive lookup
                result[str(fam).strip().lower()] = str(pl2).strip() if pl2 else ""
        wb.close()
        return result
    except Exception:
        return {}
